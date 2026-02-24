#!/usr/bin/env python3
"""Audit domain classifications by adjusted search volume.

Ranks all domains from SERP results by CTR-weighted search volume,
showing sample titles/snippets for manual review of classifications.
"""

import os
import sys
import openpyxl
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from domain_classification import classify

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEEP_FILE = os.path.join(BASE_DIR, "2. deep_parse_results.xlsx")
SERP_FILE = os.path.join(BASE_DIR, "3. serp_results.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "4. domain_audit.xlsx")

CTR_WEIGHTS = {1: 0.40, 2: 0.23, 3: 0.16, 4: 0.12, 5: 0.09}


def load_keyword_volumes():
    """Load keywords with exact match volume > 0 from deep_parse_results."""
    wb = openpyxl.load_workbook(DEEP_FILE, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    sq_col = headers.index("Search Query")
    vol_col = headers.index("Search Volume (Exact Match Type)")

    volumes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        query = str(row[sq_col]).strip() if row[sq_col] else ""
        vol = row[vol_col]
        if query and vol and float(vol) > 0:
            volumes[query] = float(vol)
    wb.close()
    print(f"Loaded {len(volumes)} keywords with exact volume > 0")
    return volumes


def resolve_domain(domain, url):
    """Resolve domain, treating yandex.ru/maps as separate from yandex.ru."""
    domain = str(domain).strip().lower() if domain else ""
    url = str(url).strip() if url else ""
    if domain == "yandex.ru" and "/maps" in url:
        return "yandex.ru/maps"
    return domain


def build_domain_data(volumes):
    """Process SERP results (positions 1-5) and aggregate by domain."""
    wb = openpyxl.load_workbook(SERP_FILE, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    sq_col = headers.index("Search Query")
    pos_col = headers.index("Position")
    url_col = headers.index("URL")
    dom_col = headers.index("Domain")
    title_col = headers.index("Title")
    snippet_col = headers.index("Snippet")

    # Per-domain accumulators
    adj_vol = defaultdict(float)
    appearances = defaultdict(int)
    titles = defaultdict(set)
    snippets = defaultdict(set)
    cat_counts = defaultdict(lambda: defaultdict(int))  # domain -> {cat -> count}

    for row in ws.iter_rows(min_row=2, values_only=True):
        query = str(row[sq_col]).strip() if row[sq_col] else ""
        pos = row[pos_col]
        if not pos or int(pos) < 1 or int(pos) > 5:
            continue
        pos = int(pos)
        if query not in volumes:
            continue

        url = str(row[url_col]).strip() if row[url_col] else ""
        raw_domain = str(row[dom_col]).strip() if row[dom_col] else ""
        domain = resolve_domain(raw_domain, url)
        if not domain:
            continue

        title = str(row[title_col]).strip() if row[title_col] else ""
        snippet = str(row[snippet_col]).strip() if row[snippet_col] else ""

        # Classify per-row (title-aware)
        if domain == "yandex.ru/maps":
            cat = classify("yandex.ru", url, title)
        else:
            cat = classify(domain, url, title)
        cat_counts[domain][cat] += 1

        weight = CTR_WEIGHTS.get(pos, 0)
        adj_vol[domain] += volumes[query] * weight
        appearances[domain] += 1

        if title and title != "None":
            titles[domain].add(title[:120])
        if snippet and snippet != "None":
            snippets[domain].add(snippet[:200])

    wb.close()

    # Build sorted list
    results = []
    for domain in sorted(adj_vol, key=lambda d: adj_vol[d], reverse=True):
        # Dominant category for this domain
        counts = cat_counts[domain]
        cat = max(counts, key=counts.get)

        # Show breakdown if mixed (e.g. "informational (85%) / commercial (15%)")
        total = sum(counts.values())
        if len(counts) > 1:
            parts = []
            for c in sorted(counts, key=counts.get, reverse=True):
                parts.append(f"{c} ({counts[c]*100//total}%)")
            cat_display = " / ".join(parts)
        else:
            cat_display = cat

        sample_titles = " | ".join(list(titles[domain])[:3])
        sample_snippets = " | ".join(list(snippets[domain])[:2])

        results.append({
            "domain": domain,
            "adj_vol": round(adj_vol[domain], 1),
            "appearances": appearances[domain],
            "category": cat_display,
            "titles": sample_titles,
            "snippets": sample_snippets,
        })

    print(f"Processed {len(results)} unique domains")
    return results


def write_output(results):
    """Write audit results to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Domain Audit"

    headers = [
        "Domain",
        "Adjusted Search Volume",
        "Appearances",
        "Current Script Category",
        "Sample Titles",
        "Sample Snippets",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r["domain"])
        ws.cell(row=i, column=2, value=r["adj_vol"])
        ws.cell(row=i, column=3, value=r["appearances"])
        ws.cell(row=i, column=4, value=r["category"])
        ws.cell(row=i, column=5, value=r["titles"])
        ws.cell(row=i, column=6, value=r["snippets"])

    wb.save(OUTPUT_FILE)
    print(f"Saved {OUTPUT_FILE}")
    print(f"Total domains: {len(results)}")


def main():
    volumes = load_keyword_volumes()
    results = build_domain_data(volumes)
    write_output(results)


if __name__ == "__main__":
    main()
