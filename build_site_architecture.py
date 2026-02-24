#!/usr/bin/env python3
"""
Build site_architecture.xlsx for Arbuzova Beauty.
Generates SEO page structure with Title, H1, Meta, internal links, queries.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict
from query_classifier import (
    load_commercial_queries, is_filtered, strip_query, categorize, SERVICE_HUBS
)

# =============================================================================
# CONFIG
# =============================================================================
XLSX_IN = '2. deep_parse_results.xlsx'
XLSX_OUT = '5. site_architecture.xlsx'
BRAND = 'Arbuzova Beauty'
CITY = 'СПб'
CITY_FULL = 'Санкт-Петербурге'
ADDRESS_SHORT = 'Комендантский пр. 53к1'
DISTRICT = 'Приморский район'

# Hub URL slugs (hub_name -> slug)
HUB_SLUGS = {
    'маникюр (главная)': '/',
    'мужской маникюр': '/muzhskoy-manikyur',
    'японский маникюр': '/yaponskiy-manikyur',
    'пилочный маникюр': '/pilochnyy-manikyur',
    'медицинский маникюр': '/meditsinskiy-manikyur',
    'маникюр и педикюр в 4 руки': '/manikyur-v-4-ruki',
    'аппаратный маникюр': '/apparatnyy-manikyur',
    'комбинированный маникюр': '/kombinirovannyy-manikyur',
    'детский маникюр': '/detskiy-manikyur',
    'гигиенический маникюр': '/gigienicheskiy-manikyur',
    'классический маникюр': '/klassicheskiy-manikyur',
    'европейский маникюр': '/evropeyskiy-manikyur',
    'маникюр с покрытием': '/manikyur-s-pokrytiem',
    'маникюр luxio': '/manikyur-luxio',
    'маникюр без покрытия': '/manikyur-bez-pokrytiya',
    'наращивание ногтей': '/narashchivanie-nogtey',
    'маникюр втирка': '/manikyur-vtirka',
    'пудровый маникюр': '/pudrovyy-manikyur',
    'маникюр омбре': '/manikyur-ombre',
    'френч маникюр': '/french-manikyur',
    'креативный маникюр': '/kreativnyy-manikyur',
    'спа маникюр': '/spa-manikyur',
    'экспресс маникюр': '/ekspress-manikyur',
    'корейский маникюр': '/koreyskiy-manikyur',
    'снятие маникюра': '/snyatie-manikyura',
    'акции и скидки на маникюр': '/aktsii-manikyur',
    'маникюр с дизайном': '/manikyur-s-dizaynom',
    'немецкий маникюр': '/nemetskiy-manikyur-lcn',
    'пленочный маникюр': '/plenochnyy-manikyur',
}

# Hub display names for titles/H1 (hub_name -> (title_name, h1_name))
HUB_NAMES = {
    'маникюр (главная)': ('Маникюр и педикюр', 'Маникюр и педикюр'),
    'мужской маникюр': ('Мужской маникюр', 'Мужской маникюр'),
    'японский маникюр': ('Японский маникюр', 'Японский маникюр'),
    'пилочный маникюр': ('Пилочный маникюр', 'Пилочный маникюр'),
    'медицинский маникюр': ('Медицинский маникюр', 'Медицинский (лечебный) маникюр'),
    'маникюр и педикюр в 4 руки': ('Маникюр и педикюр в 4 руки', 'Маникюр и педикюр в 4 руки'),
    'аппаратный маникюр': ('Аппаратный маникюр', 'Аппаратный маникюр'),
    'комбинированный маникюр': ('Комбинированный маникюр', 'Комбинированный маникюр'),
    'детский маникюр': ('Детский маникюр', 'Детский маникюр'),
    'гигиенический маникюр': ('Гигиенический маникюр', 'Гигиенический маникюр'),
    'классический маникюр': ('Классический маникюр', 'Классический (обрезной) маникюр'),
    'европейский маникюр': ('Европейский маникюр', 'Европейский (необрезной) маникюр'),
    'маникюр с покрытием': ('Маникюр с покрытием', 'Маникюр с покрытием гель-лаком'),
    'маникюр luxio': ('Маникюр Luxio', 'Маникюр с покрытием Luxio'),
    'маникюр без покрытия': ('Маникюр без покрытия', 'Маникюр без покрытия'),
    'наращивание ногтей': ('Наращивание ногтей', 'Наращивание ногтей'),
    'маникюр втирка': ('Маникюр с втиркой', 'Маникюр с втиркой'),
    'пудровый маникюр': ('Пудровый маникюр', 'Пудровый маникюр (дип-система)'),
    'маникюр омбре': ('Маникюр омбре', 'Маникюр омбре (градиент)'),
    'френч маникюр': ('Френч маникюр', 'Французский маникюр (френч)'),
    'креативный маникюр': ('Креативный маникюр', 'Креативный (арт) маникюр'),
    'спа маникюр': ('СПА маникюр', 'СПА маникюр'),
    'экспресс маникюр': ('Экспресс маникюр', 'Экспресс маникюр'),
    'корейский маникюр': ('Корейский маникюр', 'Корейский маникюр'),
    'снятие маникюра': ('Снятие маникюра', 'Снятие маникюра'),
    'акции и скидки на маникюр': ('Акции и скидки на маникюр', 'Акции и скидки на маникюр'),
    'маникюр с дизайном': ('Маникюр с дизайном', 'Маникюр с дизайном'),
    'немецкий маникюр': ('Немецкий маникюр LCN', 'Немецкий маникюр LCN'),
    'пленочный маникюр': ('Плёночный маникюр', 'Плёночный маникюр'),
}

# Meta description templates
def make_meta(hub_name):
    """Generate meta description for a hub page."""
    n = HUB_NAMES[hub_name]
    display = n[0]
    if hub_name == 'маникюр (главная)':
        return (f'{display} в {CITY_FULL} — салон {BRAND} на {ADDRESS_SHORT}, '
                f'{DISTRICT}. Запись онлайн, доступные цены, опытные мастера.')
    if hub_name == 'акции и скидки на маникюр':
        return (f'Актуальные акции и скидки на маникюр в салоне {BRAND} в {CITY}. '
                f'Выгодные предложения для новых и постоянных клиентов.')
    if hub_name == 'снятие маникюра':
        return (f'{display} гель-лаком в {CITY_FULL} — салон {BRAND}. '
                f'Бережное снятие покрытия аппаратом. Запись онлайн.')
    return (f'{display} в {CITY_FULL} — салон {BRAND} на {ADDRESS_SHORT}. '
            f'Цены, запись онлайн, опытные мастера.')


# Section assignment
def get_section(hub_name):
    if hub_name == 'маникюр (главная)':
        return 'Главная'
    if hub_name == 'акции и скидки на маникюр':
        return 'Промо'
    return 'Услуги'


# Priority assignment based on volume
def get_priority(vol):
    if vol >= 50:
        return 'High'
    if vol >= 10:
        return 'Medium'
    return 'Low'


# =============================================================================
# MAIN
# =============================================================================
def main():
    # Load and classify queries
    rows = load_commercial_queries(XLSX_IN)

    # Group queries by hub (excluding filtered and competitors)
    hub_queries = defaultdict(list)  # hub_name -> [(original_query, volume)]

    for q, v, c in rows:
        if is_filtered(q):
            continue
        stripped = strip_query(q)
        cat = categorize(stripped, q)
        if cat == 'COMPETITOR':
            continue
        hub_queries[cat].append((q, v))

    # Sort queries within each hub by volume descending
    for hub in hub_queries:
        hub_queries[hub].sort(key=lambda x: -x[1])

    # Build internal links for main page (all hub URLs)
    hub_urls = [HUB_SLUGS[h] for h in sorted(hub_queries.keys())
                if h != 'маникюр (главная)']
    main_links = ', '.join(hub_urls)

    # Build rows for XLSX
    output_rows = []

    # Main page first
    main_hub = 'маникюр (главная)'
    main_vol = sum(v for _, v in hub_queries[main_hub])
    main_queries_str = '\n'.join(f'{q} [{v}]' for q, v in hub_queries[main_hub])
    title_name = HUB_NAMES[main_hub][0]
    h1_name = HUB_NAMES[main_hub][1]

    output_rows.append({
        'Section': get_section(main_hub),
        'URL': '/',
        'Title': f'{title_name} в {CITY} | Салон {BRAND} — {ADDRESS_SHORT}',
        'H1': f'{h1_name} в {CITY_FULL} — салон {BRAND}',
        'Meta Description': make_meta(main_hub),
        'Parent Page': None,
        'Internal Links': main_links,
        'Combined Search Volume': main_vol,
        'Priority': 'High',
        'Поисковые запросы': main_queries_str,
    })

    # Hub pages sorted by volume descending
    hub_list = [(h, sum(v for _, v in qs), qs)
                for h, qs in hub_queries.items() if h != main_hub]
    hub_list.sort(key=lambda x: -x[1])

    for hub_name, total_vol, queries in hub_list:
        title_name = HUB_NAMES[hub_name][0]
        h1_name = HUB_NAMES[hub_name][1]
        url = HUB_SLUGS[hub_name]
        queries_str = '\n'.join(f'{q} [{v}]' for q, v in queries)

        # Internal links: main page + sibling hubs (top 5 by vol, excluding self)
        sibling_links = [HUB_SLUGS[h] for h, v, _ in hub_list[:6]
                        if h != hub_name and HUB_SLUGS[h] != url][:5]
        internal_links = ', '.join(['/' ] + sibling_links)

        output_rows.append({
            'Section': get_section(hub_name),
            'URL': url,
            'Title': f'{title_name} в {CITY} | Салон {BRAND}',
            'H1': f'{h1_name} в {CITY_FULL}',
            'Meta Description': make_meta(hub_name),
            'Parent Page': '/',
            'Internal Links': internal_links,
            'Combined Search Volume': total_vol,
            'Priority': get_priority(total_vol),
            'Поисковые запросы': queries_str,
        })

    # Create XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Site Architecture'

    headers = ['Section', 'URL', 'Title', 'H1', 'Meta Description',
               'Parent Page', 'Internal Links', 'Combined Search Volume',
               'Priority', 'Поисковые запросы']

    # Header styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    wrap_alignment = Alignment(vertical='top', wrap_text=True)

    for row_idx, row_data in enumerate(output_rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.alignment = wrap_alignment
            cell.border = thin_border

    # Column widths
    col_widths = {
        'A': 12,   # Section
        'B': 30,   # URL
        'C': 55,   # Title
        'D': 55,   # H1
        'E': 70,   # Meta Description
        'F': 10,   # Parent Page
        'G': 50,   # Internal Links
        'H': 15,   # Combined Search Volume
        'I': 10,   # Priority
        'J': 60,   # Поисковые запросы
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(XLSX_OUT)

    # Summary
    print(f'Created {XLSX_OUT}:')
    print(f'  {len(output_rows)} pages (1 main + {len(output_rows)-1} hubs)')
    total_vol = sum(r['Combined Search Volume'] for r in output_rows)
    total_queries = sum(len(hub_queries[h]) for h in hub_queries)
    print(f'  {total_queries} queries, combined volume = {total_vol}')

    priorities = defaultdict(int)
    for r in output_rows:
        priorities[r['Priority']] += 1
    print(f'  Priority: {dict(priorities)}')


if __name__ == '__main__':
    main()
