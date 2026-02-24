#!/usr/bin/env python3
"""
Calculate commercialization level for each keyword based on SERP results.

Weighted scoring using Backlinko CTR data for top 5 positions:
  Position 1: 40%, Position 2: 23%, Position 3: 16%, Position 4: 12%, Position 5: 9%

For each keyword, we check top 5 SERP results:
  - commercial → weight * 1.0
  - ecommerce  → weight * 0.0 (not counted)
  - informational → weight * 0.0
  - unknown → excluded from calculation (weight redistributed)

Final score = sum of commercial weights / sum of all known weights (excluding unknowns)
Result is 0.0 to 1.0 (0% to 100% commercialization).
"""

import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from domain_classification import classify

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEEP_FILE = os.path.join(BASE_DIR, "2. deep_parse_results.xlsx")
SERP_FILE = os.path.join(BASE_DIR, "3. serp_results.xlsx")

# Backlinko CTR weights by position (1-indexed)
CTR_WEIGHTS = {1: 0.40, 2: 0.23, 3: 0.16, 4: 0.12, 5: 0.09}


def main():
    # 1. Load all SERP results (top 5 positions only)
    print("Loading SERP results...")
    wb = openpyxl.load_workbook(SERP_FILE, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    sq_col = headers.index("Search Query")
    pos_col = headers.index("Position")
    url_col = headers.index("URL")
    dom_col = headers.index("Domain")
    title_col = headers.index("Title")

    # Build: query -> [(position, category), ...]
    serp_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        query = str(row[sq_col]).strip() if row[sq_col] else ""
        pos = row[pos_col]
        if not query or pos is None:
            continue
        pos = int(pos)
        if pos > 5:
            continue

        url = str(row[url_col]) if row[url_col] else ""
        domain = str(row[dom_col]).strip() if row[dom_col] else ""
        title = str(row[title_col]).strip() if row[title_col] else ""
        cat = classify(domain, url, title)

        if query not in serp_data:
            serp_data[query] = []
        serp_data[query].append((pos, cat))

    wb.close()
    print(f"  Loaded SERP data for {len(serp_data)} keywords")

    # 2. Calculate commercialization score per keyword
    scores = {}
    for query, results in serp_data.items():
        commercial_weight = 0.0
        known_weight = 0.0

        for pos, cat in results:
            w = CTR_WEIGHTS.get(pos, 0)
            if cat == "unknown":
                continue  # skip unknowns, redistribute weight
            known_weight += w
            if cat == "commercial":
                commercial_weight += w

        if known_weight > 0:
            scores[query] = round(commercial_weight / known_weight, 4)
        else:
            scores[query] = None  # all results unknown

    print(f"  Calculated scores for {len(scores)} keywords")

    # Stats
    scored = [v for v in scores.values() if v is not None]
    avg = sum(scored) / len(scored) if scored else 0
    print(f"  Average commercialization: {avg:.1%}")
    print(f"  Fully commercial (>= 0.8): {sum(1 for v in scored if v >= 0.8)}")
    print(f"  High commercial (0.5-0.8): {sum(1 for v in scored if 0.5 <= v < 0.8)}")
    print(f"  Medium (0.2-0.5): {sum(1 for v in scored if 0.2 <= v < 0.5)}")
    print(f"  Low (< 0.2): {sum(1 for v in scored if v < 0.2)}")

    # 3. Write scores to deep_parse_results.xlsx
    print(f"\nWriting scores to {DEEP_FILE}...")
    wb = openpyxl.load_workbook(DEEP_FILE)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    col_name = "Commercialization Level"
    if col_name in header_row:
        col_idx = header_row.index(col_name) + 1
    else:
        col_idx = len(header_row) + 1
        ws.cell(row=1, column=col_idx, value=col_name)

    matched = 0
    for row_num in range(2, ws.max_row + 1):
        query = ws.cell(row=row_num, column=1).value
        if query and str(query).strip() in scores:
            score = scores[str(query).strip()]
            ws.cell(row=row_num, column=col_idx, value=score)
            matched += 1

    wb.save(DEEP_FILE)
    wb.close()
    print(f"Done! Matched {matched} keywords. Column '{col_name}' written.")


if __name__ == "__main__":
    main()
