import json, urllib.request, urllib.parse, openpyxl, time, os, ssl
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

ctx = ssl.create_default_context()

# Load negative keywords
wb = openpyxl.load_workbook('/Users/egordriagin/Dev/arbuzova-beauty/wordstat_top_queries.xlsx')
ws = wb.active
neg_keywords = set()
for row in range(2, 2002):
    val = ws.cell(row=row, column=5).value
    if val and str(val).strip():
        neg_keywords.add(str(val).strip().lower())

print(f"Loaded {len(neg_keywords)} negative keywords", flush=True)

API_KEY = "bce323aab1cfc1fc05e7e7bb46cdd1c12082e84d"
API_USER = "14967"
MAX_THREADS = 10

lock = threading.Lock()
all_results = {}
completed_count = 0
error_count = 0

def fetch_wordstat(query, retries=3):
    encoded = urllib.parse.quote(query)
    url = f"https://xmlriver.com/wordstat/new/json?query={encoded}&key={API_KEY}&user={API_USER}"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=60, context=ctx) as resp:
                return json.loads(resp.read().decode('utf-8'))
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
            else:
                raise

def contains_negative(text, neg_set):
    words = text.lower().split()
    for neg in neg_set:
        neg_words = neg.split()
        if len(neg_words) == 1:
            if neg in words:
                return True
        else:
            if neg in text.lower():
                return True
    return False

def process_query(idx, query, vol, total):
    global completed_count, error_count
    try:
        data = fetch_wordstat(query)
        popular = data.get("popular", [])

        local_results = {}
        for item in popular:
            text = item["text"].lower().strip()
            v = int(item["value"])
            if not contains_negative(text, neg_keywords):
                local_results[text] = max(local_results.get(text, 0), v)

        with lock:
            for text, v in local_results.items():
                if text not in all_results or all_results[text] < v:
                    all_results[text] = v
            completed_count += 1
            if completed_count % 50 == 0:
                print(f"  Progress: {completed_count}/{total} done, {len(all_results)} unique keywords", flush=True)

        return len(popular)

    except Exception as e:
        with lock:
            error_count += 1
            completed_count += 1
        print(f"  ERROR [{idx+1}] {query}: {e}", flush=True)
        return 0

progress_file = '/Users/egordriagin/Dev/arbuzova-beauty/deep_parse_progress.json'
seed_cache = '/Users/egordriagin/Dev/arbuzova-beauty/seed_cache.json'

# Step 1: Get seed results
if os.path.exists(seed_cache):
    with open(seed_cache, 'r') as f:
        seed_kept = json.load(f)
    print(f"Loaded {len(seed_kept)} cached seed keywords", flush=True)
else:
    print("Fetching seed: маникюр", flush=True)
    seed_data = fetch_wordstat("маникюр")
    seed_popular = seed_data.get("popular", [])
    print(f"Seed returned {len(seed_popular)} keywords", flush=True)

    seed_kept = []
    for item in seed_popular:
        text = item["text"]
        volume = int(item["value"])
        if not contains_negative(text, neg_keywords):
            seed_kept.append([text, volume])

    with open(seed_cache, 'w') as f:
        json.dump(seed_kept, f, ensure_ascii=False)
    print(f"After filtering: {len(seed_kept)} keywords to query deeper", flush=True)

# Add seed results first
for text, vol in seed_kept:
    key = text.lower().strip()
    all_results[key] = max(all_results.get(key, 0), vol)

# Check for resume
start_idx = 0
if os.path.exists(progress_file):
    with open(progress_file, 'r') as f:
        saved = json.load(f)
        if saved.get('last_idx', 0) >= len(seed_kept) - 1:
            # Already completed, just rebuild Excel
            all_results = saved['results']
            start_idx = len(seed_kept)
            print(f"Already completed! {len(all_results)} results loaded.", flush=True)
        else:
            all_results.update(saved.get('results', {}))
            start_idx = saved.get('last_idx', 0) + 1
            print(f"Resuming from index {start_idx}, {len(all_results)} results so far", flush=True)

total = len(seed_kept)
queries_to_run = [(i, q, v) for i, (q, v) in enumerate(seed_kept) if i >= start_idx]

if queries_to_run:
    print(f"\nStarting {len(queries_to_run)} API calls with {MAX_THREADS} threads...", flush=True)
    t_start = time.time()

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = {
            executor.submit(process_query, idx, query, vol, total): (idx, query)
            for idx, query, vol in queries_to_run
        }

        for future in as_completed(futures):
            pass  # Results handled inside process_query

    elapsed = time.time() - t_start
    print(f"\nAll queries done in {elapsed:.0f}s ({elapsed/60:.1f} min)", flush=True)
    print(f"Errors: {error_count}", flush=True)

# Save progress
with open(progress_file, 'w') as f:
    json.dump({'results': all_results, 'last_idx': total - 1}, f, ensure_ascii=False)

# Step 3: Write Excel
print(f"\nTotal unique keywords: {len(all_results)}", flush=True)
print("Writing to Excel...", flush=True)

sorted_results = sorted(all_results.items(), key=lambda x: -x[1])

out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "Deep Parse Results"
out_ws.cell(row=1, column=1, value="Search Query")
out_ws.cell(row=1, column=2, value="Search Volume")

from openpyxl.styles import Font
for col in [1, 2]:
    out_ws.cell(row=1, column=col).font = Font(bold=True)

for idx, (query, volume) in enumerate(sorted_results, start=2):
    out_ws.cell(row=idx, column=1, value=query)
    out_ws.cell(row=idx, column=2, value=volume)

out_ws.column_dimensions['A'].width = 60
out_ws.column_dimensions['B'].width = 15
out_ws.auto_filter.ref = f"A1:B{len(sorted_results)+1}"

out_wb.save('/Users/egordriagin/Dev/arbuzova-beauty/deep_parse_results.xlsx')
print(f"DONE! Saved {len(sorted_results)} rows to deep_parse_results.xlsx", flush=True)
