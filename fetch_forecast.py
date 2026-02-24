#!/usr/bin/env python3
"""
Fetch exact match search volume (Shows) for all keywords via Yandex Direct Forecast API.
Converts each keyword to exact match format: "!word1 !word2 !word3"
Region: St. Petersburg (GeoID=2).
Results are saved incrementally, then merged into deep_parse_results.xlsx.
"""

import json
import os
import re
import time
import urllib.request
import urllib.error
from datetime import datetime

import openpyxl

# ── Config ──────────────────────────────────────────────────────────────────
OAUTH_TOKEN = "y0__xC6paeDCBjskz4ggbqmxBYw5YydsQiz6WLkUoB0rcY0mp0gl9yR0MkQrQ"
API_URL = "https://api.direct.yandex.com/live/v4/json/"
REGION_IDS = [2]  # St. Petersburg
CURRENCY = "RUB"

BATCH_SIZE = 100
POLL_INTERVAL = 10
MAX_POLL_ATTEMPTS = 60  # 10 minutes

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, "2. deep_parse_results.xlsx")
PROGRESS_FILE = os.path.join(BASE_DIR, "forecast_progress.json")


# ── Exact match conversion ──────────────────────────────────────────────────

def to_exact_match(phrase: str) -> str:
    """Convert 'маникюр на короткие ногти' → '"!маникюр !на !короткие !ногти"'"""
    words = phrase.strip().split()
    exact_words = " ".join(f"!{w}" for w in words)
    return f'"{exact_words}"'


# ── API calls ───────────────────────────────────────────────────────────────

def api_call(method: str, param=None, max_retries: int = 3) -> dict:
    body = {
        "method": method,
        "locale": "ru",
        "token": OAUTH_TOKEN,
    }
    if param is not None:
        body["param"] = param

    data = json.dumps(body, ensure_ascii=False).encode("utf-8")

    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(
                API_URL, data=data,
                headers={"Content-Type": "application/json; charset=utf-8"},
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json.loads(resp.read().decode("utf-8"))

            if "error_code" in result:
                code = result["error_code"]
                msg = result.get("error_detail", result.get("error_str", ""))

                if code == 31:
                    return {"error_code": 31}
                elif code == 152:
                    print(f"    Low points, waiting 60s...")
                    time.sleep(60)
                    continue
                elif code in (52, 53):
                    time.sleep(5 if code == 53 else 60)
                    continue
                elif code >= 1000:
                    print(f"    AUTH ERROR ({code}): {msg}")
                    raise SystemExit(1)
                elif code in (500, 502, 503):
                    time.sleep(2 ** (attempt + 1))
                    continue
                else:
                    print(f"    API error {code}: {msg}")
                    return result

            return result

        except (urllib.error.HTTPError, urllib.error.URLError, OSError) as e:
            time.sleep(2 ** (attempt + 1))
        except SystemExit:
            raise
        except Exception as e:
            time.sleep(2 ** (attempt + 1))

    return {"error_code": -1}


def create_forecast(phrases: list[str]) -> int | None:
    result = api_call("CreateNewForecast", {
        "Phrases": phrases,
        "GeoID": REGION_IDS,
        "Currency": CURRENCY,
    })
    if "data" in result:
        return result["data"]
    if result.get("error_code") == 31:
        cleanup_reports()
        result = api_call("CreateNewForecast", {
            "Phrases": phrases,
            "GeoID": REGION_IDS,
            "Currency": CURRENCY,
        })
        if "data" in result:
            return result["data"]
    return None


def wait_for_forecast(forecast_id: int) -> bool:
    for _ in range(MAX_POLL_ATTEMPTS):
        result = api_call("GetForecastList")
        for r in result.get("data", []):
            if r.get("ForecastID") == forecast_id:
                status = r.get("StatusForecast", "")
                if status == "Done":
                    return True
                if status == "Failed":
                    print(f"    Forecast {forecast_id} FAILED")
                    return False
                break
        time.sleep(POLL_INTERVAL)
    print(f"    Forecast {forecast_id} TIMEOUT")
    return False


def get_forecast(forecast_id: int) -> list[dict]:
    result = api_call("GetForecast", forecast_id)
    if "data" not in result:
        return []
    return result["data"].get("Phrases", [])


def delete_forecast(forecast_id: int):
    api_call("DeleteForecastReport", forecast_id)


def cleanup_reports():
    result = api_call("GetForecastList")
    for r in result.get("data", []):
        fid = r.get("ForecastID")
        if fid:
            delete_forecast(fid)
    time.sleep(2)


# ── Progress tracking ───────────────────────────────────────────────────────

def load_progress() -> dict:
    """Load {original_phrase: shows} from progress file."""
    if not os.path.exists(PROGRESS_FILE):
        return {}
    with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_progress(progress: dict):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False)


# ── Main ────────────────────────────────────────────────────────────────────

def load_keywords() -> list[tuple]:
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    ws = wb.active
    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        query, volume, category = row[0], row[1], row[2]
        if query:
            keywords.append((str(query).strip(), volume, category))
    wb.close()
    return keywords


def merge_into_source(progress: dict):
    """Add 'Search Volume (Exact)' column to deep_parse_results.xlsx."""
    print(f"\nMerging {len(progress)} results into {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # Find or create the column
    header_row = [cell.value for cell in ws[1]]
    col_name = "Search Volume (Exact)"
    if col_name in header_row:
        col_idx = header_row.index(col_name) + 1
    else:
        col_idx = len(header_row) + 1
        ws.cell(row=1, column=col_idx, value=col_name)

    # Fill in values
    matched = 0
    for row_num in range(2, ws.max_row + 1):
        query = ws.cell(row=row_num, column=1).value
        if query and str(query).strip() in progress:
            shows = progress[str(query).strip()]
            ws.cell(row=row_num, column=col_idx, value=shows)
            matched += 1

    wb.save(INPUT_FILE)
    wb.close()
    print(f"Done! Matched {matched} keywords. Column '{col_name}' added to {INPUT_FILE}")


def main():
    keywords = load_keywords()
    total = len(keywords)
    print(f"Loaded {total} keywords from {INPUT_FILE}")

    progress = load_progress()
    print(f"Already completed: {len(progress)} keywords")

    remaining = [(q, v, c) for q, v, c in keywords if q not in progress]
    print(f"Remaining: {len(remaining)} keywords")

    batches = [remaining[i:i + BATCH_SIZE] for i in range(0, len(remaining), BATCH_SIZE)]
    total_batches = len(batches)
    print(f"Batches: {total_batches} (×{BATCH_SIZE})")
    print(f"Started: {datetime.now().strftime('%H:%M:%S')}")
    print()

    for batch_idx, batch in enumerate(batches):
        batch_num = batch_idx + 1
        done_count = len(progress)
        pct = done_count / total * 100

        # Build mapping: exact_match_phrase → original_phrase
        exact_to_original = {}
        exact_phrases = []
        for q, v, c in batch:
            exact = to_exact_match(q)
            exact_to_original[exact] = q
            exact_phrases.append(exact)

        print(f"[{batch_num}/{total_batches}] ({pct:.1f}%) {len(batch)} phrases")

        # Create forecast
        forecast_id = create_forecast(exact_phrases)
        if forecast_id is None:
            print(f"    Failed to create forecast, skipping batch")
            for q, v, c in batch:
                progress[q] = None
            save_progress(progress)
            continue

        # Wait
        if not wait_for_forecast(forecast_id):
            delete_forecast(forecast_id)
            for q, v, c in batch:
                progress[q] = None
            save_progress(progress)
            continue

        # Get results
        phrases_data = get_forecast(forecast_id)
        delete_forecast(forecast_id)

        # Map results back to original phrases
        for item in phrases_data:
            api_phrase = item.get("Phrase", "")
            original = exact_to_original.get(api_phrase)
            if original:
                progress[original] = item.get("Shows", 0)
            else:
                # Try fuzzy match — API may normalize the phrase
                for exact, orig in exact_to_original.items():
                    if orig not in progress:
                        # Compare lowercased without operators
                        api_clean = re.sub(r'["!]', '', api_phrase).strip().lower()
                        orig_clean = orig.strip().lower()
                        if api_clean == orig_clean:
                            progress[orig] = item.get("Shows", 0)
                            break

        # Mark any unmatched as 0
        for q, v, c in batch:
            if q not in progress:
                progress[q] = 0

        save_progress(progress)
        print(
            f"    Done: {len(phrases_data)} results | "
            f"Total: {len(progress)}/{total} | "
            f"{datetime.now().strftime('%H:%M:%S')}"
        )

    # Final merge into source file
    merge_into_source(progress)

    print(f"\nAll done! {datetime.now().strftime('%H:%M:%S')}")


if __name__ == "__main__":
    main()
