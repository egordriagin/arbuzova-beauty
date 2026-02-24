#!/usr/bin/env python3
"""
Fetch top 10 Yandex SERP results for all keywords via XMLRiver API.
Region: St. Petersburg (lr=2). Results saved to serp_results.xlsx.
Supports resume — reads existing progress from serp_results.xlsx.
Uses 10 parallel workers for speed.
"""

import os
import re
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from threading import Lock

import openpyxl

# ── Config ──────────────────────────────────────────────────────────────────
API_USER = "14967"
API_KEY = "bce323aab1cfc1fc05e7e7bb46cdd1c12082e84d"
LR = 2  # St. Petersburg
GROUPS_ON_PAGE = 10
DOCS_IN_GROUP = 1
WORKERS = 10
SAVE_EVERY = 100  # flush to disk every N completed queries

INPUT_FILE = "deep_parse_results.xlsx"
OUTPUT_FILE = "serp_results.xlsx"

HEADERS = [
    "Search Query",
    "Search Volume",
    "Категория",
    "Position",
    "URL",
    "Domain",
    "Title",
    "Snippet",
    "Cache URL",
]

# ── Helpers ─────────────────────────────────────────────────────────────────

def strip_hlword(text: str) -> str:
    """Remove <hlword>...</hlword> tags, keeping inner text."""
    if not text:
        return ""
    text = re.sub(r"</?hlword>", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def element_text_raw(el) -> str:
    """Get the full inner XML text of an element (including child tags) as a string."""
    if el is None:
        return ""
    raw = ET.tostring(el, encoding="unicode", method="xml")
    inner = re.sub(r"^<[^>]+>", "", raw, count=1)
    inner = re.sub(r"</[^>]+>$", "", inner, count=1)
    return inner.strip()


def build_url(query: str) -> str:
    encoded_query = urllib.parse.quote(query)
    groupby = (
        f"attr%3Dd.mode%3Ddeep"
        f".groups-on-page%3D{GROUPS_ON_PAGE}"
        f".docs-in-group%3D{DOCS_IN_GROUP}"
    )
    return (
        f"http://xmlriver.com/yandex/xml"
        f"?user={API_USER}"
        f"&key={API_KEY}"
        f"&query={encoded_query}"
        f"&lr={LR}"
        f"&groupby={groupby}"
    )


def fetch_serp(query: str, max_retries: int = 3) -> list[dict]:
    """Fetch top 10 SERP results for a query. Returns list of dicts."""
    url = build_url(query)

    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = resp.read().decode("utf-8")

            root = ET.fromstring(data)

            error = root.find(".//error")
            if error is not None:
                error_code = error.get("code", "")
                error_text = error.text or ""
                if error_code == "500" or "перезапрос" in error_text.lower():
                    time.sleep(2 * (attempt + 1))
                    continue
                else:
                    return []

            results = []
            groups = root.findall(".//group")
            for pos, group in enumerate(groups, start=1):
                doc = group.find("doc")
                if doc is None:
                    continue

                url_el = doc.find("url")
                domain_el = doc.find("domain")
                title_el = doc.find("title")
                cache_el = doc.find("saved-copy-url")

                passages = doc.findall(".//passage")
                snippet_parts = [element_text_raw(p) for p in passages]
                snippet = " ".join(snippet_parts)
                if not snippet:
                    headline_el = doc.find("headline")
                    snippet = element_text_raw(headline_el)

                results.append({
                    "position": pos,
                    "url": url_el.text if url_el is not None else "",
                    "domain": domain_el.text if domain_el is not None else "",
                    "title": strip_hlword(element_text_raw(title_el)),
                    "snippet": strip_hlword(snippet),
                    "cache_url": cache_el.text if cache_el is not None else "",
                })

            return results

        except Exception:
            time.sleep(2 * (attempt + 1))

    return []


def make_rows(query, volume, category, results: list[dict]) -> list:
    if results:
        return [
            [query, volume, category,
             r["position"], r["url"], r["domain"],
             r["title"], r["snippet"], r["cache_url"]]
            for r in results
        ]
    else:
        return [[query, volume, category, None, "NO_RESULTS", None, None, None, None]]


# ── I/O ─────────────────────────────────────────────────────────────────────

def load_keywords() -> list[tuple]:
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    ws = wb.active
    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        query, volume, category = row[0], row[1], row[2]
        if query:
            keywords.append((str(query).strip(), volume, category))
    wb.close()
    return keywords


def load_completed_queries() -> set:
    if not os.path.exists(OUTPUT_FILE):
        return set()
    wb = openpyxl.load_workbook(OUTPUT_FILE, read_only=True)
    ws = wb.active
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            done.add(str(row[0]).strip())
    wb.close()
    return done


def init_output():
    if os.path.exists(OUTPUT_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SERP Results"
    ws.append(HEADERS)
    wb.save(OUTPUT_FILE)
    wb.close()


def save_batch(buffer: list):
    if not buffer:
        return
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    for row in buffer:
        ws.append(row)
    wb.save(OUTPUT_FILE)
    wb.close()


# ── Main ────────────────────────────────────────────────────────────────────

def process_keyword(item):
    """Worker function: fetch SERP for one keyword, return rows."""
    query, volume, category = item
    results = fetch_serp(query)
    return query, make_rows(query, volume, category, results)


def main():
    keywords = load_keywords()
    total = len(keywords)
    print(f"Loaded {total} keywords from {INPUT_FILE}")

    init_output()
    done = load_completed_queries()
    print(f"Already completed: {len(done)} keywords")

    remaining = [(q, v, c) for q, v, c in keywords if q not in done]
    print(f"Remaining: {len(remaining)} keywords")
    print(f"Workers: {WORKERS}")
    print(f"Started at {datetime.now().strftime('%H:%M:%S')}")
    print()

    if not remaining:
        print("Nothing to do!")
        return

    buffer = []
    completed = 0
    errors = 0
    lock = Lock()

    with ThreadPoolExecutor(max_workers=WORKERS) as executor:
        futures = {
            executor.submit(process_keyword, item): item[0]
            for item in remaining
        }

        for future in as_completed(futures):
            query = futures[future]
            try:
                query, rows = future.result()
                is_error = len(rows) == 1 and rows[0][4] == "NO_RESULTS"
            except Exception as e:
                rows = [[query, None, None, None, "ERROR", None, str(e), None, None]]
                is_error = True

            with lock:
                buffer.extend(rows)
                completed += 1
                if is_error:
                    errors += 1

                total_done = len(done) + completed
                pct = total_done / total * 100

                if completed % 10 == 0 or completed == len(remaining):
                    print(
                        f"[{total_done}/{total}] ({pct:.1f}%) "
                        f"+{completed} done, {errors} errors  "
                        f"| {datetime.now().strftime('%H:%M:%S')}"
                    )

                if completed % SAVE_EVERY == 0:
                    save_batch(buffer)
                    buffer.clear()
                    print(f"  ── saved to disk ({total_done} queries total) ──")

    # Final flush
    save_batch(buffer)
    print()
    print(f"Done! {len(done) + completed} queries total ({errors} errors)")
    print(f"Results saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
