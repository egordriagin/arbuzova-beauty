#!/usr/bin/env python3
"""
Scan SERP results for keywords sorted by exact search volume,
classify domains, and report unknowns for batch classification.
Processes keywords from position START to END (1-indexed).
"""

import os
import sys
from collections import Counter

import openpyxl

# Add project dir to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from domain_classification import classify

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEEP_FILE = os.path.join(BASE_DIR, "2. deep_parse_results.xlsx")
SERP_FILE = os.path.join(BASE_DIR, "3. serp_results.xlsx")

START = int(sys.argv[1]) if len(sys.argv) > 1 else 301
END = int(sys.argv[2]) if len(sys.argv) > 2 else 2000


def main():
    # 1. Load keywords with exact volumes, sort by volume desc
    print("Loading keywords...")
    wb = openpyxl.load_workbook(DEEP_FILE, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Find columns
    query_col = 0
    vol_exact_col = None
    for i, h in enumerate(headers):
        if h and "exact" in str(h).lower():
            vol_exact_col = i
            break

    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        q = row[0]
        vol = row[vol_exact_col] if vol_exact_col is not None else 0
        if q:
            keywords.append((str(q).strip(), vol if vol else 0))
    wb.close()

    # Sort by exact volume descending
    keywords.sort(key=lambda x: -(x[1] if isinstance(x[1], (int, float)) else 0))

    # Get slice
    selected = keywords[START - 1 : END]
    selected_queries = {q for q, v in selected}
    print(f"Keywords {START}-{END}: {len(selected)} keywords")
    print(f"Top volume: {selected[0][1] if selected else 'N/A'}, Bottom volume: {selected[-1][1] if selected else 'N/A'}")

    # 2. Load SERP results for these keywords (top 5 positions only)
    print("Loading SERP results...")
    wb = openpyxl.load_workbook(SERP_FILE, read_only=True)
    ws = wb.active
    serp_headers = [cell.value for cell in ws[1]]

    # Find columns in SERP file
    sq_col = serp_headers.index("Search Query") if "Search Query" in serp_headers else 0
    pos_col = serp_headers.index("Position") if "Position" in serp_headers else 3
    url_col = serp_headers.index("URL") if "URL" in serp_headers else 4
    dom_col = serp_headers.index("Domain") if "Domain" in serp_headers else 5
    title_col = serp_headers.index("Title") if "Title" in serp_headers else 6

    unknown_domains = Counter()
    classified_counts = Counter()
    unknown_examples = {}  # domain -> list of (query, position, url)

    for row in ws.iter_rows(min_row=2, values_only=True):
        query = str(row[sq_col]).strip() if row[sq_col] else ""
        if query not in selected_queries:
            continue
        pos = row[pos_col]
        if pos is None or (isinstance(pos, (int, float)) and pos > 5):
            continue

        url = str(row[url_col]) if row[url_col] else ""
        domain = str(row[dom_col]).strip() if row[dom_col] else ""

        title = str(row[title_col]).strip() if row[title_col] else ""
        cat = classify(domain, url, title)
        classified_counts[cat] += 1

        if cat == "unknown":
            unknown_domains[domain] += 1
            if domain not in unknown_examples:
                unknown_examples[domain] = []
            if len(unknown_examples[domain]) < 2:
                unknown_examples[domain].append((query, pos, url))

    wb.close()

    # 3. Report
    print(f"\nClassification summary (top 5 positions):")
    for cat, cnt in sorted(classified_counts.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {cnt}")

    total_classified = sum(classified_counts.values())
    unknown_count = classified_counts.get("unknown", 0)
    known_pct = (1 - unknown_count / total_classified) * 100 if total_classified else 0
    print(f"\nKnown: {known_pct:.1f}% ({total_classified - unknown_count}/{total_classified})")

    print(f"\n{'='*80}")
    print(f"Unknown domains ({len(unknown_domains)} unique, {unknown_count} total):")
    print(f"{'='*80}")
    for domain, count in unknown_domains.most_common():
        examples = unknown_examples.get(domain, [])
        ex_str = " | ".join(f"'{q}' pos={p}" for q, p, u in examples[:2])
        print(f"  {domain} ({count}x) — {ex_str}")


if __name__ == "__main__":
    main()
