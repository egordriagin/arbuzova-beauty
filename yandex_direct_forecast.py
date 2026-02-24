#!/usr/bin/env python3
"""
Yandex Direct API v4 (Live) — Forecast: получение прогноза показов по ключевым фразам.

Для каждой фразы возвращается Shows (прогноз показов за месяц) в регионе Санкт-Петербург.
Результаты сохраняются в forecast_results.xlsx. Поддержка возобновления при сбое.

Использование:
    python3 yandex_direct_forecast.py

Перед запуском укажите OAuth-токен в переменной OAUTH_TOKEN.
"""

import json
import os
import time
import urllib.request
import urllib.error
from datetime import datetime

import openpyxl

# ── Конфигурация ────────────────────────────────────────────────────────────
OAUTH_TOKEN = "y0__xC6paeDCBjskz4ggbqmxBYw5YydsQiz6WLkUoB0rcY0mp0gl9yR0MkQrQ"

API_URL = "https://api.direct.yandex.com/live/v4/json/"
REGION_IDS = [2]  # Санкт-Петербург
CURRENCY = "RUB"

MAX_PHRASES_PER_FORECAST = 100   # лимит API
MAX_CONCURRENT_REPORTS = 5       # максимум отчётов на сервере
POLL_INTERVAL = 15               # секунды между проверками статуса
MAX_POLL_ATTEMPTS = 40           # ~10 минут ожидания на отчёт

INPUT_FILE = "deep_parse_results.xlsx"
OUTPUT_FILE = "forecast_results.xlsx"

OUTPUT_HEADERS = [
    "Ключевая фраза",
    "Частотность (Wordstat)",
    "Категория",
    "Прогноз показов (Shows)",
    "Клики",
    "Клики (1-е место)",
    "Клики (спецразмещение)",
    "CTR",
    "CTR (1-е место)",
    "CTR (спецразмещение)",
    "Мин. цена клика",
    "Макс. цена клика",
    "Мин. цена (спецразмещение)",
    "Макс. цена (спецразмещение)",
]

# ── Вызовы API ──────────────────────────────────────────────────────────────

def api_call(method: str, params=None, max_retries: int = 3) -> dict:
    """
    Универсальный вызов Yandex Direct API v4 (Live).
    """
    body = {
        "method": method,
        "locale": "ru",
        "token": OAUTH_TOKEN,
    }
    if params is not None:
        body["param"] = params

    data = json.dumps(body, ensure_ascii=False).encode("utf-8")

    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(
                API_URL,
                data=data,
                headers={"Content-Type": "application/json; charset=utf-8"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                response_data = resp.read().decode("utf-8")
                result = json.loads(response_data)

            # Проверка ошибок API
            if "error_code" in result:
                code = result["error_code"]
                msg = result.get("error_detail", result.get("error_str", ""))

                if code == 31:
                    # Превышен лимит отчётов — нужно удалить старые
                    print(f"    Лимит отчётов (код 31), очистка...")
                    return {"error_code": 31, "error": msg}
                elif code == 152:
                    # Нехватка баллов
                    print(f"    Нехватка баллов (код 152), пауза 60 сек...")
                    time.sleep(60)
                    continue
                elif code in (52, 53):
                    delay = 5 if code == 53 else 60
                    print(f"    Ошибка {code}, пауза {delay} сек...")
                    time.sleep(delay)
                    continue
                elif code in (500, 502, 503, 1000, 1001, 1002):
                    if code >= 1000:
                        print(f"    КРИТИЧЕСКАЯ ОШИБКА авторизации ({code}): {msg}")
                        raise SystemExit(1)
                    delay = 2 ** (attempt + 1)
                    print(f"    Ошибка сервера {code}, повтор через {delay} сек...")
                    time.sleep(delay)
                    continue
                else:
                    print(f"    Ошибка API {code}: {msg}")
                    return result

            return result

        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8") if e.fp else ""
            delay = 2 ** (attempt + 1)
            print(f"    HTTP {e.code}, повтор через {delay} сек ({attempt+1}/{max_retries})")
            time.sleep(delay)

        except Exception as e:
            delay = 2 ** (attempt + 1)
            print(f"    Ошибка: {e}, повтор через {delay} сек ({attempt+1}/{max_retries})")
            time.sleep(delay)

    print(f"    СБОЙ после {max_retries} попыток")
    return {"error_code": -1, "error": "max_retries_exceeded"}


def create_forecast(phrases: list[str]) -> int | None:
    """Создаёт прогноз, возвращает ForecastID или None."""
    result = api_call("CreateNewForecast", {
        "Phrases": phrases,
        "GeoID": REGION_IDS,
        "Currency": CURRENCY,
    })

    if "data" in result:
        return result["data"]

    # Если лимит отчётов — очищаем и повторяем
    if result.get("error_code") == 31:
        cleanup_reports()
        result = api_call("CreateNewForecast", {
            "Phrases": phrases,
            "GeoID": REGION_IDS,
            "Currency": CURRENCY,
        })
        if "data" in result:
            return result["data"]

    return None


def wait_for_forecast(forecast_id: int) -> bool:
    """Ожидает завершения отчёта. Возвращает True если Done."""
    for _ in range(MAX_POLL_ATTEMPTS):
        result = api_call("GetForecastList")
        reports = result.get("data", [])

        for r in reports:
            if r.get("ForecastID") == forecast_id:
                status = r.get("StatusForecast", "")
                if status == "Done":
                    return True
                elif status == "Failed":
                    print(f"    Отчёт {forecast_id} — ошибка генерации")
                    return False
                break

        time.sleep(POLL_INTERVAL)

    print(f"    Таймаут ожидания отчёта {forecast_id}")
    return False


def get_forecast(forecast_id: int) -> list[dict]:
    """Получает данные прогноза. Возвращает список фраз с показателями."""
    result = api_call("GetForecast", forecast_id)

    if "data" not in result:
        return []

    phrases = result["data"].get("Phrases", [])
    return phrases


def delete_forecast(forecast_id: int):
    """Удаляет отчёт с сервера."""
    api_call("DeleteForecastReport", forecast_id)


def cleanup_reports():
    """Удаляет все существующие отчёты."""
    result = api_call("GetForecastList")
    reports = result.get("data", [])
    for r in reports:
        fid = r.get("ForecastID")
        if fid:
            delete_forecast(fid)
            print(f"    Удалён отчёт {fid}")
    time.sleep(2)


# ── Работа с файлами ────────────────────────────────────────────────────────

def load_keywords() -> list[tuple]:
    """Загрузка (фраза, частотность, категория) из входного файла."""
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    ws = wb.active
    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        query, volume, category = row[0], row[1], row[2]
        if query:
            keywords.append((str(query).strip(), volume, category))
    wb.close()
    return keywords


def load_completed_keywords() -> set:
    """Загрузка уже обработанных фраз."""
    if not os.path.exists(OUTPUT_FILE):
        return set()
    wb = openpyxl.load_workbook(OUTPUT_FILE, read_only=True)
    ws = wb.active
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            done.add(str(row[0]).strip())
    wb.close()
    return done


def init_output():
    """Создание выходного файла с заголовками."""
    if os.path.exists(OUTPUT_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast Results"
    ws.append(OUTPUT_HEADERS)
    wb.save(OUTPUT_FILE)
    wb.close()


def save_batch(rows: list):
    """Сохранение батча результатов."""
    if not rows:
        return
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(OUTPUT_FILE)
    wb.close()


# ── Основной цикл ──────────────────────────────────────────────────────────

def main():
    if OAUTH_TOKEN == "ВСТАВЬТЕ_ВАШ_OAUTH_ТОКЕН":
        print("ОШИБКА: Укажите OAuth-токен в переменной OAUTH_TOKEN")
        return

    keywords = load_keywords()
    total = len(keywords)
    print(f"Загружено {total} фраз из {INPUT_FILE}")

    init_output()
    done = load_completed_keywords()
    print(f"Уже обработано: {len(done)} фраз")

    remaining = [(q, v, c) for q, v, c in keywords if q not in done]
    print(f"Осталось: {len(remaining)} фраз")

    # Разбивка на батчи по 100 фраз
    batches = []
    for i in range(0, len(remaining), MAX_PHRASES_PER_FORECAST):
        batches.append(remaining[i : i + MAX_PHRASES_PER_FORECAST])

    total_batches = len(batches)
    print(f"Батчей: {total_batches} (по {MAX_PHRASES_PER_FORECAST} фраз)")
    print(f"Старт: {datetime.now().strftime('%H:%M:%S')}")
    print()

    for batch_idx, batch in enumerate(batches):
        batch_num = batch_idx + 1
        kw_meta = {q: (v, c) for q, v, c in batch}
        kw_list = [q for q, v, c in batch]

        pct = (len(done) + batch_idx * MAX_PHRASES_PER_FORECAST) / total * 100
        print(f"[Батч {batch_num}/{total_batches}] ({pct:.1f}%) {len(batch)} фраз")

        # 1. Создание прогноза
        forecast_id = create_forecast(kw_list)
        if forecast_id is None:
            print(f"    Не удалось создать прогноз, пропускаем батч")
            error_rows = [
                [q, v, c, "ERROR", None, None, None, None, None, None, None, None, None, None]
                for q, (v, c) in kw_meta.items()
            ]
            save_batch(error_rows)
            continue

        print(f"    Создан прогноз ID={forecast_id}, ожидание...")

        # 2. Ожидание готовности
        if not wait_for_forecast(forecast_id):
            print(f"    Прогноз не готов, пропускаем батч")
            delete_forecast(forecast_id)
            error_rows = [
                [q, v, c, "TIMEOUT", None, None, None, None, None, None, None, None, None, None]
                for q, (v, c) in kw_meta.items()
            ]
            save_batch(error_rows)
            continue

        # 3. Получение результатов
        phrases_data = get_forecast(forecast_id)

        # 4. Удаление отчёта (освобождаем слот)
        delete_forecast(forecast_id)

        # 5. Парсинг и сохранение
        output_rows = []
        for item in phrases_data:
            phrase = item.get("Phrase", "")
            volume, category = kw_meta.get(phrase, (None, None))
            output_rows.append([
                phrase,
                volume,
                category,
                item.get("Shows"),
                item.get("Clicks"),
                item.get("FirstPlaceClicks"),
                item.get("PremiumClicks"),
                item.get("CTR"),
                item.get("FirstPlaceCTR"),
                item.get("PremiumCTR"),
                item.get("Min"),
                item.get("Max"),
                item.get("PremiumMin"),
                item.get("PremiumMax"),
            ])

        save_batch(output_rows)
        print(
            f"    Готово: {len(output_rows)} фраз сохранено "
            f"| {datetime.now().strftime('%H:%M:%S')}"
        )

    print()
    print(f"Завершено! Результаты: {OUTPUT_FILE}")
    print(f"Время: {datetime.now().strftime('%H:%M:%S')}")


if __name__ == "__main__":
    main()
