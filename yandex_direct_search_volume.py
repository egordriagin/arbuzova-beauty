#!/usr/bin/env python3
"""
Yandex Direct API — KeywordsResearch.hasSearchVolume

Проверка наличия поискового объёма для ключевых фраз в регионе Санкт-Петербург.
Результаты сохраняются в XLSX-файл с разбивкой по типам устройств.

Использование:
    python3 yandex_direct_search_volume.py

Перед запуском укажите OAuth-токен в переменной OAUTH_TOKEN.
"""

import json
import os
import time
import urllib.request
import urllib.error
from datetime import datetime

import openpyxl

# ── Конфигурация ────────────────────────────────────────────────────────────
OAUTH_TOKEN = "ВСТАВЬТЕ_ВАШ_OAUTH_ТОКЕН"

API_URL = "https://api.direct.yandex.com/json/v5/keywordsresearch"
REGION_IDS = [2]  # Санкт-Петербург

# Лимиты API
MAX_KEYWORDS_PER_REQUEST = 10000  # максимум фраз в одном запросе
MAX_REQUESTS_PER_MINUTE = 20     # максимум запросов за 60 секунд
REQUEST_INTERVAL = 60.0 / MAX_REQUESTS_PER_MINUTE  # 3 секунды между запросами

INPUT_FILE = "deep_parse_results.xlsx"
OUTPUT_FILE = "search_volume_results.xlsx"

OUTPUT_HEADERS = [
    "Ключевая фраза",
    "Частотность (Wordstat)",
    "Категория",
    "Все устройства",
    "Десктопы",
    "Мобильные",
    "Планшеты",
]

# ── Работа с API ────────────────────────────────────────────────────────────

def build_request_body(keywords: list[str]) -> str:
    """Формирование JSON-тела запроса к hasSearchVolume."""
    body = {
        "method": "hasSearchVolume",
        "params": {
            "SelectionCriteria": {
                "Keywords": keywords,
                "RegionIds": REGION_IDS,
            },
            "FieldNames": [
                "Keyword",
                "RegionIds",
                "AllDevices",
                "MobilePhones",
                "Tablets",
                "Desktops",
            ],
        },
    }
    return json.dumps(body, ensure_ascii=False)


def call_api(request_body: str, max_retries: int = 3) -> dict:
    """
    Отправка POST-запроса к API Яндекс.Директа.

    Обрабатывает:
      - HTTP 200: успешный ответ
      - HTTP 500/502/503: повторный запрос с экспоненциальной задержкой
      - Ошибки баллов (Units): пауза до восстановления
    """
    headers = {
        "Authorization": f"Bearer {OAUTH_TOKEN}",
        "Accept-Language": "ru",
        "Content-Type": "application/json; charset=utf-8",
    }

    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(
                API_URL,
                data=request_body.encode("utf-8"),
                headers=headers,
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                # Чтение заголовка Units (израсходовано/остаток/лимит)
                units_header = resp.getheader("Units", "")
                if units_header:
                    parts = units_header.split("/")
                    if len(parts) == 3:
                        spent, remaining, daily_limit = parts
                        print(
                            f"    Баллы: -{spent}, "
                            f"остаток {remaining}/{daily_limit}"
                        )
                        # Пауза при низком остатке баллов
                        if int(remaining) < 100:
                            print(
                                "    ⚠ Мало баллов, пауза 60 секунд..."
                            )
                            time.sleep(60)

                response_data = resp.read().decode("utf-8")
                return json.loads(response_data)

        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8") if e.fp else ""

            if e.code in (500, 502, 503):
                # Внутренняя ошибка сервера — повтор с задержкой
                delay = 2 ** (attempt + 1)
                print(
                    f"    Ошибка HTTP {e.code}, "
                    f"повтор через {delay} сек "
                    f"({attempt + 1}/{max_retries})"
                )
                time.sleep(delay)
                continue

            elif e.code == 400:
                # Ошибка в параметрах запроса
                error_info = json.loads(error_body) if error_body else {}
                error_code = error_info.get("error", {}).get("error_code", "")

                if str(error_code) == "52":
                    # NOT_ENOUGH_UNITS — нехватка баллов
                    print("    ⚠ Нехватка баллов, пауза 60 секунд...")
                    time.sleep(60)
                    continue
                elif str(error_code) == "53":
                    # CONCURRENT_LIMIT — превышен лимит соединений
                    print(
                        "    ⚠ Лимит соединений, пауза 5 секунд..."
                    )
                    time.sleep(5)
                    continue
                elif str(error_code) == "506":
                    # TOO_MANY_OBJECTS — слишком много объектов
                    print(f"    Ошибка: слишком много фраз в запросе")
                    return {"error": error_info}
                else:
                    print(f"    Ошибка API {error_code}: {error_body}")
                    return {"error": error_info}

            elif e.code in (1000, 1001, 1002):
                # Ошибки авторизации — остановка
                print(f"    КРИТИЧЕСКАЯ ОШИБКА авторизации (HTTP {e.code})")
                print(f"    {error_body}")
                raise SystemExit(1)

            else:
                print(f"    Ошибка HTTP {e.code}: {error_body}")
                return {"error": error_body}

        except Exception as e:
            delay = 2 ** (attempt + 1)
            print(
                f"    Ошибка сети: {e}, "
                f"повтор через {delay} сек "
                f"({attempt + 1}/{max_retries})"
            )
            time.sleep(delay)

    print(f"    СБОЙ после {max_retries} попыток")
    return {"error": "max_retries_exceeded"}


# ── Работа с файлами ────────────────────────────────────────────────────────

def load_keywords() -> list[tuple]:
    """Загрузка ключевых фраз из входного файла."""
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    ws = wb.active
    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        query, volume, category = row[0], row[1], row[2]
        if query:
            keywords.append((str(query).strip(), volume, category))
    wb.close()
    return keywords


def load_completed_keywords() -> set:
    """Загрузка уже обработанных фраз из файла результатов."""
    if not os.path.exists(OUTPUT_FILE):
        return set()
    wb = openpyxl.load_workbook(OUTPUT_FILE, read_only=True)
    ws = wb.active
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            done.add(str(row[0]).strip())
    wb.close()
    return done


def init_output():
    """Создание выходного файла с заголовками."""
    if os.path.exists(OUTPUT_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Search Volume"
    ws.append(OUTPUT_HEADERS)
    wb.save(OUTPUT_FILE)
    wb.close()


def save_batch(rows: list):
    """Сохранение батча результатов в файл."""
    if not rows:
        return
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(OUTPUT_FILE)
    wb.close()


# ── Основной цикл ──────────────────────────────────────────────────────────

def main():
    # Проверка токена
    if OAUTH_TOKEN == "ВСТАВЬТЕ_ВАШ_OAUTH_ТОКЕН":
        print("ОШИБКА: Укажите OAuth-токен в переменной OAUTH_TOKEN")
        print("Получить токен: https://oauth.yandex.ru/authorize?response_type=token&client_id=ВАШ_CLIENT_ID")
        return

    # Загрузка данных
    keywords = load_keywords()
    total = len(keywords)
    print(f"Загружено {total} фраз из {INPUT_FILE}")

    init_output()
    done = load_completed_keywords()
    print(f"Уже обработано: {len(done)} фраз")

    remaining = [(q, v, c) for q, v, c in keywords if q not in done]
    print(f"Осталось обработать: {len(remaining)} фраз")

    # Разбивка на батчи
    batches = []
    for i in range(0, len(remaining), MAX_KEYWORDS_PER_REQUEST):
        batches.append(remaining[i : i + MAX_KEYWORDS_PER_REQUEST])

    total_batches = len(batches)
    print(f"Батчей: {total_batches} (по {MAX_KEYWORDS_PER_REQUEST} фраз)")
    print(f"Лимит: {MAX_REQUESTS_PER_MINUTE} запросов/мин")
    print(f"Старт: {datetime.now().strftime('%H:%M:%S')}")
    print()

    for batch_idx, batch in enumerate(batches):
        batch_num = batch_idx + 1
        print(f"[Батч {batch_num}/{total_batches}] {len(batch)} фраз")

        # Словарь для быстрого поиска volume/category по фразе
        kw_meta = {q: (v, c) for q, v, c in batch}
        kw_list = [q for q, v, c in batch]

        # Запрос к API
        body = build_request_body(kw_list)
        response = call_api(body)

        # Обработка ошибки
        if "error" in response:
            print(f"    Ошибка в батче {batch_num}, пропускаем")
            # Сохраняем фразы с пометкой об ошибке
            error_rows = [
                [q, v, c, "ERROR", "ERROR", "ERROR", "ERROR"]
                for q, (v, c) in kw_meta.items()
            ]
            save_batch(error_rows)
            time.sleep(REQUEST_INTERVAL)
            continue

        # Парсинг результатов
        results = response.get("result", {}).get("HasSearchVolumeResults", [])
        output_rows = []

        for item in results:
            keyword = item.get("Keyword", "")
            volume, category = kw_meta.get(keyword, (None, None))
            output_rows.append([
                keyword,
                volume,
                category,
                item.get("AllDevices", ""),
                item.get("Desktops", ""),
                item.get("MobilePhones", ""),
                item.get("Tablets", ""),
            ])

        # Сохранение батча
        save_batch(output_rows)
        processed = len(done) + (batch_idx + 1) * MAX_KEYWORDS_PER_REQUEST
        processed = min(processed, total)
        pct = processed / total * 100
        print(
            f"    Сохранено. "
            f"Прогресс: {processed}/{total} ({pct:.1f}%) "
            f"| {datetime.now().strftime('%H:%M:%S')}"
        )

        # Пауза между запросами (соблюдение лимита 20 запросов/мин)
        if batch_idx < total_batches - 1:
            time.sleep(REQUEST_INTERVAL)

    print()
    print(f"Готово! Результаты: {OUTPUT_FILE}")
    print(f"Завершено: {datetime.now().strftime('%H:%M:%S')}")


if __name__ == "__main__":
    main()
